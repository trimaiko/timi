import logging
from datetime import date, datetime, timedelta
from openpyxl import load_workbook


class timemanager:

    def __init__(self, data, outputfile):
        # data is python object
        self.data = data
        self.outputfile = outputfile

        self.logger = logging.getLogger(__name__)
        self.break_hour = "1:00"

    def edit_timecard(self):
        self.logger.info("edit_timecard start")

        base_row = 8
        arrive_base_col = 6
        depart_base_col = 7
        break_base_col = 8
        remote_base_col = 3

        wb = load_workbook(filename = self.outputfile)
        #always select first sheet in workbook
        ws = wb["就業記録簿"]

        # self.logger.info(self.data)

        # insert arrival time
        for dat in self.data['data']:
            if dat.get('ユーザーID') == '':
                continue

            mycell_arrive = ws.cell(row = base_row, column= arrive_base_col)
            mycell_dept = ws.cell(row = base_row, column = depart_base_col)
            mycell_break = ws.cell(row = base_row, column = break_base_col)
            mycell_remote = ws.cell(row = base_row, column = remote_base_col)

            if dat.get('出社時刻') != "":
                self.logger.info(dat.get('日付'))
                self.logger.info(dat.get('出社時刻'))
                atime = datetime.strptime(dat.get('出社時刻'), '%H:%M')
                # mycell_arrive.value = self.convert_date_to_excel_ordinal(atime)
                mycell_arrive.value = atime - timedelta(days=1)
                mycell_arrive.number_format = "[h]:mm"

                self.logger.info('mycell_arrive.value:', mycell_arrive.value)
                self.logger.info('atime:', atime)
                self.logger.info(dat.get('出社時刻'))

                dtime = datetime.strptime(dat.get('退社時刻'), '%H:%M')
                self.logger.info(dat.get('退社時刻'))
                mycell_dept.value = dtime - timedelta(days=1)
                mycell_dept.number_format = "[h]:mm"
                print('mycell_dept.value:', mycell_dept.value)
                print('dtime:', dtime)

                if dtime - atime > timedelta(hours=6):
                    mycell_break.value = datetime.strptime(self.break_hour, '%H:%M') - timedelta(days=1)
                # read remark
                remark = dat.get("備考")
                if "リモート" in remark:
                    mycell_remote.value = "リモート"
                elif "出張" in remark:
                    mycell_remote.value = "出張"
                    self.logger.info(mycell_remote)
            else:
                mycell_arrive.value = ""
                mycell_dept.value = ""


            base_row += 1

        # write_to_outputfile

        wb.save(filename = self.outputfile)